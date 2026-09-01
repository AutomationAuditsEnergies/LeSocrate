import io
import unittest
from datetime import date, datetime, timedelta, timezone
from unittest.mock import patch

from openpyxl import load_workbook

from services import attendance_service


class AttendanceConsolidationTest(unittest.TestCase):
    def test_dashboard_keeps_enrolled_student_absent_without_heartbeat(self):
        course_session = {
            "id": 42,
            "scheduled_at": datetime(2026, 7, 17, 9, 0, tzinfo=timezone.utc),
            "timezone": "Europe/Paris",
        }
        platform = {
            "id": 3,
            "center_platform_number": 1,
            "name": "TP EC",
            "center_account_id": 9,
            "source_module_id": 77,
        }
        with patch.object(attendance_service.attendance_repo, "get_accessible_platform", return_value=platform), patch.object(
            attendance_service.attendance_repo, "get_course_session_for_date", return_value=course_session
        ), patch.object(
            attendance_service.attendance_repo, "list_presence_logs_for_session", return_value=[]
        ), patch.object(
            attendance_service, "list_explicit_course_reminder_recipients",
            return_value=[{"id": 8, "email": "lina@example.test", "nom": "Martin", "prenom": "Lina"}],
        ), patch.object(attendance_service.attendance_repo, "list_daily_exports", return_value=[]):
            payload = attendance_service.get_attendance_dashboard(3, "2026-07-17")

        self.assertEqual(len(payload["students"]), 1)
        self.assertEqual(payload["students"][0]["id"], "recipient:8")
        self.assertEqual(payload["students"][0]["attendance"]["status"], "absent")

    def test_reconnects_are_kept_and_overlapping_tabs_are_not_double_counted(self):
        paris = timezone(timedelta(hours=2))
        scheduled_at = datetime(2026, 7, 17, 9, 0, tzinfo=paris)
        rows = [
            {
                "recipient_hash": "a" * 64,
                "nom": "Dupuis",
                "prenom": "Thomas",
                "email": "thomas@example.test",
                "attendance_started_at": datetime(2026, 7, 17, 8, 55, tzinfo=paris),
                "depart": datetime(2026, 7, 17, 10, 0, tzinfo=paris),
            },
            {
                "recipient_hash": "a" * 64,
                "nom": "Dupuis",
                "prenom": "Thomas",
                "email": "thomas@example.test",
                "attendance_started_at": datetime(2026, 7, 17, 9, 30, tzinfo=paris),
                "depart": datetime(2026, 7, 17, 10, 30, tzinfo=paris),
            },
            {
                "recipient_hash": "a" * 64,
                "nom": "Dupuis",
                "prenom": "Thomas",
                "email": "thomas@example.test",
                "attendance_started_at": datetime(2026, 7, 17, 11, 0, tzinfo=paris),
                "last_seen_at": datetime(2026, 7, 17, 12, 0, tzinfo=paris),
            },
        ]

        participants = attendance_service.consolidate_presence(
            rows,
            scheduled_at=scheduled_at,
        )

        self.assertEqual(len(participants), 1)
        self.assertEqual(len(participants[0]["intervals"]), 2)
        self.assertEqual(participants[0]["intervals"][0][0].hour, 9)
        self.assertEqual(participants[0]["total_seconds"], 2 * 60 * 60 + 30 * 60)

    def test_presence_is_limited_to_training_day_and_excludes_lunch(self):
        paris = timezone(timedelta(hours=2))
        scheduled_at = datetime(2026, 7, 17, 9, 0, tzinfo=paris)
        rows = [{
            "recipient_hash": "c" * 64,
            "nom": "Martin",
            "prenom": "Lina",
            "email": "lina@example.test",
            "attendance_started_at": datetime(2026, 7, 17, 8, 45, tzinfo=paris),
            "depart": datetime(2026, 7, 17, 17, 30, tzinfo=paris),
        }]
        schedule_blocks = [
            {"block_type": "course", "start_minute": 540, "end_minute": 720},
            {
                "block_type": "pause",
                "pause_kind": "lunch",
                "start_minute": 720,
                "end_minute": 780,
            },
            {"block_type": "course", "start_minute": 780, "end_minute": 1020},
        ]

        participants = attendance_service.consolidate_presence(
            rows,
            scheduled_at=scheduled_at,
            schedule_blocks=schedule_blocks,
        )

        self.assertEqual(len(participants[0]["intervals"]), 2)
        self.assertEqual(participants[0]["intervals"][0][0].hour, 9)
        self.assertEqual(participants[0]["intervals"][0][1].hour, 12)
        self.assertEqual(participants[0]["intervals"][1][0].hour, 13)
        self.assertEqual(participants[0]["intervals"][1][1].hour, 17)
        self.assertEqual(participants[0]["total_seconds"], 7 * 60 * 60)

    def test_daily_workbook_contains_summary_detail_and_auditable_formulas(self):
        paris = timezone(timedelta(hours=2))
        participants = [{
            "key": "invite:" + "b" * 64,
            "nom": "Dupuis",
            "prenom": "Thomas",
            "email": "thomas@example.test",
            "intervals": [
                (
                    datetime(2026, 7, 17, 9, 2, tzinfo=paris),
                    datetime(2026, 7, 17, 10, 14, tzinfo=paris),
                ),
                (
                    datetime(2026, 7, 17, 10, 20, tzinfo=paris),
                    datetime(2026, 7, 17, 12, 0, tzinfo=paris),
                ),
            ],
            "total_seconds": 10560,
        }]

        payload = attendance_service.generate_daily_attendance_excel(
            center_name="Centre Horizon",
            center_account_id=9,
            platform_name="TP CRCD",
            platform_id=42,
            center_platform_number=1,
            course_session_id=103,
            teacher_module_id=77,
            course_date=date(2026, 7, 17),
            session_index=3,
            participants=participants,
        )
        workbook = load_workbook(io.BytesIO(payload), data_only=False)

        self.assertEqual(workbook.sheetnames, ["Synthèse", "Détail connexions"])
        self.assertIn("center_account_id=9", workbook.properties.keywords)
        self.assertIn("course_session_id=103", workbook.properties.keywords)
        summary = workbook["Synthèse"]
        details = workbook["Détail connexions"]
        self.assertEqual(summary["B5"].value, "Dupuis")
        self.assertEqual(summary["C5"].value, "Thomas")
        self.assertIn("Centre Horizon · Plateforme 1", summary["A2"].value)
        self.assertIn("SUMIF", summary["G5"].value)
        self.assertEqual(details["B5"].value, "Dupuis")
        self.assertEqual(details["H5"].value, "=F5-E5")
        self.assertTrue(summary.column_dimensions["A"].hidden)


class AttendanceSchedulerTest(unittest.TestCase):
    def test_due_export_is_claimed_published_and_completed_once(self):
        now = datetime(2026, 7, 18, 4, 5, tzinfo=timezone.utc)
        job = {
            "id": 17,
            "platform_id": 3,
            "course_session_id": 42,
            "course_date": date(2026, 7, 17),
        }
        session = {
            "id": 42,
            "platform_id": 3,
            "center_account_id": 9,
            "center_platform_number": 1,
            "center_name": "Centre Horizon",
            "teacher_module_id": 77,
            "session_index": 2,
            "scheduled_at": datetime(2026, 7, 17, 7, 0, tzinfo=timezone.utc),
            "platform_name": "TP CRCD",
            "timezone": "Europe/Paris",
        }
        with patch.object(attendance_service, "schedule_store_is_postgres", return_value=True), patch.object(
            attendance_service.attendance_repo, "close_stale_presence_logs", return_value=1
        ), patch.object(
            attendance_service.attendance_repo, "materialize_daily_export_candidates", return_value=1
        ), patch.object(
            attendance_service.attendance_repo, "claim_due_daily_export", side_effect=[job, None]
        ), patch.object(
            attendance_service.attendance_repo, "get_course_session", return_value=session
        ), patch.object(
            attendance_service.attendance_repo, "list_presence_logs_for_session", return_value=[]
        ), patch.object(
            attendance_service, "publish_daily_attendance_excel",
            return_value={"container": "formation-attendance", "blob_key": "p3/j42/file.xlsx", "size": 1234},
        ) as publish, patch.object(
            attendance_service.attendance_repo, "complete_daily_export", return_value=True
        ) as complete:
            results = attendance_service.process_due_attendance_exports(now=now)

        self.assertEqual(results, [{"success": True, "export_id": 17, "participant_count": 0}])
        publish.assert_called_once()
        self.assertEqual(publish.call_args.kwargs["center_account_id"], 9)
        self.assertEqual(publish.call_args.kwargs["center_platform_number"], 1)
        complete.assert_called_once()
        self.assertEqual(complete.call_args.args[0], 17)

    def test_blob_key_is_scoped_by_center_and_local_platform_number(self):
        key = attendance_service.attendance_blob_key(
            9,
            1,
            42,
            103,
            "presences-2026-07-17.xlsx",
        )

        self.assertEqual(
            key,
            "centres/9/plateformes/1/id-42/seances/103/presences/presences-2026-07-17.xlsx",
        )


if __name__ == "__main__":
    unittest.main()
