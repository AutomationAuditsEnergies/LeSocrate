"""Automatic room attendance, daily Excel generation and persistent storage."""

from __future__ import annotations

from datetime import datetime, timedelta, timezone
from hashlib import sha256
import io
import os
import re
from typing import Any
from zoneinfo import ZoneInfo

from azure.core.exceptions import ResourceExistsError
from azure.storage.blob import BlobServiceClient, ContentSettings
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from repositories import attendance_repository as attendance_repo
from repositories.course_schedule_repository import (
    list_explicit_course_reminder_recipients,
    schedule_store_is_postgres,
)
from utils.auth_tokens import course_invitation_recipient_hash
from utils.logger import get_logger
from utils.slug import slugify


logger = get_logger(__name__)
ATTENDANCE_CONTAINER = "formation-attendance"
HEARTBEAT_STALE_SECONDS = 90


def _as_datetime(value, tz_name: str = "Europe/Paris") -> datetime | None:
    if not value:
        return None
    if isinstance(value, datetime):
        parsed = value
    else:
        parsed = datetime.fromisoformat(str(value).strip().replace("Z", "+00:00"))
    zone = ZoneInfo(tz_name)
    if parsed.tzinfo is None:
        return parsed.replace(tzinfo=zone)
    return parsed.astimezone(zone)


def _participant_key(row: dict[str, Any]) -> str:
    recipient_hash = str(row.get("recipient_hash") or "").strip()
    if recipient_hash:
        return f"invite:{recipient_hash}"
    normalized = "|".join(
        re.sub(r"\s+", " ", str(row.get(field) or "").strip().lower())
        for field in ("nom", "prenom")
    )
    return f"name:{normalized}"


def _merge_intervals(intervals: list[tuple[datetime, datetime]]) -> list[tuple[datetime, datetime]]:
    merged: list[list[datetime]] = []
    for start, end in sorted(intervals, key=lambda item: item[0]):
        if end <= start:
            continue
        if not merged or start > merged[-1][1]:
            merged.append([start, end])
        elif end > merged[-1][1]:
            merged[-1][1] = end
    return [(item[0], item[1]) for item in merged]


def consolidate_presence(
    rows: list[dict[str, Any]],
    *,
    scheduled_at,
    timezone_name: str = "Europe/Paris",
) -> list[dict[str, Any]]:
    """Consolidate reconnects and overlapping tabs into auditable intervals."""
    course_start = _as_datetime(scheduled_at, timezone_name)
    grouped: dict[str, dict[str, Any]] = {}
    for row in rows:
        started = _as_datetime(row.get("attendance_started_at"), timezone_name)
        ended = _as_datetime(row.get("depart") or row.get("last_seen_at"), timezone_name)
        if not started or not ended:
            continue
        if course_start:
            started = max(started, course_start)
        if ended <= started:
            continue
        key = _participant_key(row)
        participant = grouped.setdefault(
            key,
            {
                "key": key,
                "nom": str(row.get("nom") or "").strip(),
                "prenom": str(row.get("prenom") or "").strip(),
                "email": str(row.get("email") or "").strip(),
                "intervals": [],
            },
        )
        if not participant["email"] and row.get("email"):
            participant["email"] = str(row["email"]).strip()
        participant["intervals"].append((started, ended))

    participants = []
    for participant in grouped.values():
        intervals = _merge_intervals(participant["intervals"])
        total_seconds = int(sum((end - start).total_seconds() for start, end in intervals))
        participants.append({
            **participant,
            "intervals": intervals,
            "total_seconds": max(0, total_seconds),
        })
    return sorted(
        participants,
        key=lambda item: (item["nom"].casefold(), item["prenom"].casefold(), item["email"].casefold()),
    )


def _duration_label(seconds: int) -> str:
    seconds = max(0, int(seconds or 0))
    hours, remainder = divmod(seconds, 3600)
    minutes, secs = divmod(remainder, 60)
    if hours:
        return f"{hours} h {minutes:02d} min {secs:02d} s"
    return f"{minutes} min {secs:02d} s"


def build_daily_attendance_workbook(
    *,
    center_name: str = "",
    center_account_id: int | None = None,
    platform_name: str,
    platform_id: int | None = None,
    center_platform_number: int | None = None,
    course_session_id: int | None = None,
    teacher_module_id: int | None = None,
    course_date,
    session_index: int,
    participants: list[dict[str, Any]],
) -> Workbook:
    """Create the two-sheet daily report used as the legal attendance trace."""
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Synthèse"
    details = workbook.create_sheet("Détail connexions")
    workbook.properties.title = f"Présences {platform_name} {course_date}"
    workbook.properties.subject = "Relevé quotidien des présences"
    workbook.properties.creator = "Le Socrate"
    workbook.properties.keywords = ";".join(
        f"{key}={value}"
        for key, value in (
            ("center_account_id", center_account_id),
            ("center_platform_number", center_platform_number),
            ("platform_id", platform_id),
            ("course_session_id", course_session_id),
            ("teacher_module_id", teacher_module_id),
        )
        if value is not None
    )

    violet = "7C3AED"
    ink = "172033"
    slate = "516078"
    line = "DDE3EC"
    white = "FFFFFF"
    thin = Side(style="thin", color=line)

    max_intervals = max((len(item["intervals"]) for item in participants), default=1)
    summary_headers = [
        "Clé du relevé",
        "Nom",
        "Prénom",
        "E-mail invité",
        "Première entrée",
        "Dernière sortie",
        "Temps total",
    ]
    for index in range(1, max_intervals + 1):
        summary_headers.extend([f"Entrée {index}", f"Sortie {index}"])

    last_summary_col = get_column_letter(len(summary_headers))
    summary.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(summary_headers))
    summary["A1"] = f"Présences · {platform_name}"
    summary["A1"].font = Font(name="Aptos Display", size=18, bold=True, color=ink)
    summary["A1"].alignment = Alignment(vertical="center")
    summary.row_dimensions[1].height = 30
    summary.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(summary_headers))
    owner_label = str(center_name or "Centre de formation").strip()
    local_platform_label = (
        f"Plateforme {int(center_platform_number)}"
        if center_platform_number is not None
        else "Plateforme"
    )
    summary["A2"] = (
        f"{owner_label} · {local_platform_label} · Journée {int(session_index or 0)}"
        f" · {course_date.strftime('%d/%m/%Y')} · {len(participants)} participant(s)"
    )
    summary["A2"].font = Font(name="Aptos", size=10, color=slate)
    summary["A2"].alignment = Alignment(vertical="center")

    for col, title in enumerate(summary_headers, start=1):
        cell = summary.cell(row=4, column=col, value=title)
        cell.fill = PatternFill("solid", fgColor=violet)
        cell.font = Font(name="Aptos", size=10, bold=True, color=white)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    summary.row_dimensions[4].height = 30

    report_key_by_participant = {
        participant["key"]: f"P{index:04d}"
        for index, participant in enumerate(participants, start=1)
    }
    detail_rows: list[tuple[str, str, datetime, datetime, int]] = []
    for participant in participants:
        for start, end in participant["intervals"]:
            detail_rows.append((
                report_key_by_participant[participant["key"]],
                participant["key"],
                start,
                end,
                int((end - start).total_seconds()),
            ))

    for row_index, participant in enumerate(participants, start=5):
        intervals = participant["intervals"]
        values: list[Any] = [
            report_key_by_participant[participant["key"]],
            participant["nom"],
            participant["prenom"],
            participant["email"],
            intervals[0][0].replace(tzinfo=None) if intervals else None,
            intervals[-1][1].replace(tzinfo=None) if intervals else None,
            None,
        ]
        for interval_index in range(max_intervals):
            if interval_index < len(intervals):
                start, end = intervals[interval_index]
                values.extend([start.replace(tzinfo=None), end.replace(tzinfo=None)])
            else:
                values.extend([None, None])
        for col, value in enumerate(values, start=1):
            summary.cell(row=row_index, column=col, value=value)
        detail_last_row = max(4, 4 + len(detail_rows))
        summary.cell(row=row_index, column=7, value=(
            f"=SUMIF('Détail connexions'!$A$5:$A${detail_last_row},"
            f"$A{row_index},'Détail connexions'!$H$5:$H${detail_last_row})"
        ))
        for col in range(5, len(summary_headers) + 1):
            summary.cell(row=row_index, column=col).number_format = (
                "[h]:mm:ss" if col == 7 else "dd/mm/yyyy hh:mm:ss"
            )
        if row_index % 2 == 1:
            for col in range(1, len(summary_headers) + 1):
                summary.cell(row=row_index, column=col).fill = PatternFill("solid", fgColor="FAFBFD")
        for col in range(1, len(summary_headers) + 1):
            summary.cell(row=row_index, column=col).border = Border(bottom=thin)

    if participants:
        table = Table(displayName="SynthesePresences", ref=f"A4:{last_summary_col}{4 + len(participants)}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium4",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=False,
            showColumnStripes=False,
        )
        summary.add_table(table)
    else:
        summary.merge_cells(start_row=5, start_column=2, end_row=6, end_column=min(7, len(summary_headers)))
        summary["B5"] = "Aucune entrée dans la salle n’a été enregistrée pour cette journée."
        summary["B5"].font = Font(name="Aptos", size=11, color=slate, italic=True)
        summary["B5"].alignment = Alignment(vertical="center", wrap_text=True)

    summary.column_dimensions["A"].hidden = True
    widths = {"B": 20, "C": 20, "D": 32, "E": 21, "F": 21, "G": 16}
    for letter, width in widths.items():
        summary.column_dimensions[letter].width = width
    for col in range(8, len(summary_headers) + 1):
        summary.column_dimensions[get_column_letter(col)].width = 21
    summary.freeze_panes = "B5"
    summary.auto_filter.ref = f"B4:{last_summary_col}{max(5, 4 + len(participants))}"
    summary.sheet_view.showGridLines = False
    summary.sheet_properties.pageSetUpPr.fitToPage = True
    summary.page_setup.fitToWidth = 1
    summary.page_setup.fitToHeight = 0
    summary.print_title_rows = "1:4"

    detail_headers = [
        "Clé du relevé",
        "Nom",
        "Prénom",
        "E-mail invité",
        "Entrée",
        "Sortie",
        "Durée lisible",
        "Durée calculée",
    ]
    details.merge_cells("A1:H1")
    details["A1"] = f"Détail des connexions · {platform_name}"
    details["A1"].font = Font(name="Aptos Display", size=16, bold=True, color=ink)
    details.merge_cells("A2:H2")
    details["A2"] = (
        f"{owner_label} · {local_platform_label} · Journée du "
        f"{course_date.strftime('%d/%m/%Y')} · horaires Europe/Paris"
    )
    details["A2"].font = Font(name="Aptos", size=10, color=slate)
    for col, title in enumerate(detail_headers, start=1):
        cell = details.cell(row=4, column=col, value=title)
        cell.fill = PatternFill("solid", fgColor=ink)
        cell.font = Font(name="Aptos", size=10, bold=True, color=white)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    participant_by_key = {item["key"]: item for item in participants}
    for row_index, (report_key, participant_key, start, end, seconds) in enumerate(detail_rows, start=5):
        participant = participant_by_key[participant_key]
        details.cell(row=row_index, column=1, value=report_key)
        details.cell(row=row_index, column=2, value=participant["nom"])
        details.cell(row=row_index, column=3, value=participant["prenom"])
        details.cell(row=row_index, column=4, value=participant["email"])
        details.cell(row=row_index, column=5, value=start.replace(tzinfo=None))
        details.cell(row=row_index, column=6, value=end.replace(tzinfo=None))
        details.cell(row=row_index, column=7, value=_duration_label(seconds))
        details.cell(row=row_index, column=8, value=f"=F{row_index}-E{row_index}")
        details.cell(row=row_index, column=5).number_format = "dd/mm/yyyy hh:mm:ss"
        details.cell(row=row_index, column=6).number_format = "dd/mm/yyyy hh:mm:ss"
        details.cell(row=row_index, column=8).number_format = "[h]:mm:ss"
        for col in range(1, 9):
            details.cell(row=row_index, column=col).border = Border(bottom=thin)

    if detail_rows:
        table = Table(displayName="DetailConnexions", ref=f"A4:H{4 + len(detail_rows)}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        details.add_table(table)
    details.column_dimensions["A"].hidden = True
    for letter, width in {"B": 20, "C": 20, "D": 32, "E": 24, "F": 24, "G": 21, "H": 18}.items():
        details.column_dimensions[letter].width = width
    details.freeze_panes = "B5"
    details.sheet_view.showGridLines = False
    details.sheet_properties.pageSetUpPr.fitToPage = True
    details.page_setup.fitToWidth = 1
    details.page_setup.fitToHeight = 0
    details.print_title_rows = "1:4"
    return workbook


def generate_daily_attendance_excel(
    *,
    center_name: str = "",
    center_account_id: int | None = None,
    platform_name: str,
    platform_id: int | None = None,
    center_platform_number: int | None = None,
    course_session_id: int | None = None,
    teacher_module_id: int | None = None,
    course_date,
    session_index: int,
    participants: list[dict[str, Any]],
) -> bytes:
    workbook = build_daily_attendance_workbook(
        center_name=center_name,
        center_account_id=center_account_id,
        platform_name=platform_name,
        platform_id=platform_id,
        center_platform_number=center_platform_number,
        course_session_id=course_session_id,
        teacher_module_id=teacher_module_id,
        course_date=course_date,
        session_index=session_index,
        participants=participants,
    )
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _storage_connection_string() -> str:
    value = (os.environ.get("AZURE_STORAGE_CONNECTION_STRING") or "").strip()
    if not value:
        raise ValueError("Connexion Azure des relevés de présence manquante")
    return value


def attendance_container() -> str:
    value = (os.environ.get("AZURE_ATTENDANCE_CONTAINER") or ATTENDANCE_CONTAINER).strip().lower()
    if not re.fullmatch(r"[a-z0-9](?:[a-z0-9-]{1,61}[a-z0-9])", value):
        raise ValueError("Nom du conteneur de présence invalide")
    return value


def attendance_blob_key(
    center_account_id: int,
    center_platform_number: int,
    platform_id: int,
    session_id: int,
    filename: str,
) -> str:
    """Return the immutable tenant-owned location of one attendance file."""
    return (
        f"centres/{int(center_account_id)}/"
        f"plateformes/{int(center_platform_number)}/"
        f"id-{int(platform_id)}/seances/{int(session_id)}/presences/{filename}"
    )


def publish_daily_attendance_excel(
    *,
    center_account_id: int,
    center_platform_number: int,
    platform_id: int,
    session_id: int,
    filename: str,
    excel_bytes: bytes,
    blob_service_client=None,
) -> dict[str, Any]:
    client = blob_service_client or BlobServiceClient.from_connection_string(_storage_connection_string())
    container_name = attendance_container()
    container = client.get_container_client(container_name)
    try:
        container.create_container()
    except ResourceExistsError:
        pass
    blob_key = attendance_blob_key(
        center_account_id,
        center_platform_number,
        platform_id,
        session_id,
        filename,
    )
    container.get_blob_client(blob_key).upload_blob(
        excel_bytes,
        overwrite=True,
        content_settings=ContentSettings(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            content_disposition=f'attachment; filename="{filename}"',
        ),
    )
    return {"container": container_name, "blob_key": blob_key, "size": len(excel_bytes)}


def download_daily_attendance_excel(export_row: dict[str, Any], *, blob_service_client=None) -> bytes:
    client = blob_service_client or BlobServiceClient.from_connection_string(_storage_connection_string())
    blob = client.get_blob_client(
        container=str(export_row["container_name"]),
        blob=str(export_row["blob_key"]),
    )
    return blob.download_blob().readall()


def process_due_attendance_exports(*, now: datetime | None = None, max_exports: int = 8) -> list[dict[str, Any]]:
    if not schedule_store_is_postgres():
        return []
    now = now or datetime.now(timezone.utc)
    closed = attendance_repo.close_stale_presence_logs(
        cutoff=now - timedelta(seconds=HEARTBEAT_STALE_SECONDS),
    )
    materialized = attendance_repo.materialize_daily_export_candidates(now=now)
    results: list[dict[str, Any]] = []
    for _ in range(max(1, min(50, int(max_exports)))):
        job = attendance_repo.claim_due_daily_export(now=now)
        if not job:
            break
        export_id = int(job["id"])
        try:
            session = attendance_repo.get_course_session(int(job["course_session_id"]))
            if not session:
                raise RuntimeError("Séance de formation introuvable")
            rows = attendance_repo.list_presence_logs_for_session(
                int(session["platform_id"]),
                int(session["id"]),
            )
            participants = consolidate_presence(
                rows,
                scheduled_at=session["scheduled_at"],
                timezone_name=session.get("timezone") or "Europe/Paris",
            )
            course_date = job["course_date"]
            platform_slug = slugify(
                session["platform_name"],
                fallback=f"formation-{int(session['platform_id'])}",
            )
            filename = (
                f"presences-{platform_slug}"
                f"-{course_date.isoformat()}.xlsx"
            )
            excel_bytes = generate_daily_attendance_excel(
                center_name=session.get("center_name") or "",
                center_account_id=int(session["center_account_id"]),
                platform_name=session["platform_name"],
                platform_id=int(session["platform_id"]),
                center_platform_number=int(session["center_platform_number"]),
                course_session_id=int(session["id"]),
                teacher_module_id=(
                    int(session["teacher_module_id"])
                    if session.get("teacher_module_id") is not None
                    else None
                ),
                course_date=course_date,
                session_index=int(session.get("session_index") or 0),
                participants=participants,
            )
            published = publish_daily_attendance_excel(
                center_account_id=int(session["center_account_id"]),
                center_platform_number=int(session["center_platform_number"]),
                platform_id=int(session["platform_id"]),
                session_id=int(session["id"]),
                filename=filename,
                excel_bytes=excel_bytes,
            )
            completed_at = datetime.now(timezone.utc)
            completed = attendance_repo.complete_daily_export(
                export_id,
                now=completed_at,
                container_name=published["container"],
                blob_key=published["blob_key"],
                filename=filename,
                size_bytes=published["size"],
                sha256=sha256(excel_bytes).hexdigest(),
                participant_count=len(participants),
            )
            if not completed:
                raise RuntimeError("Le verrou de génération du relevé a expiré")
            results.append({"success": True, "export_id": export_id, "participant_count": len(participants)})
        except Exception as exc:
            logger.exception("ATTENDANCE_EXPORT_FAILED export_id=%s", export_id)
            failure = attendance_repo.fail_daily_export(
                export_id,
                now=datetime.now(timezone.utc),
                error=f"{type(exc).__name__}: {str(exc)[:800]}",
            )
            results.append({"success": False, "export_id": export_id, "retry": failure})
    if closed or materialized or results:
        logger.info(
            "ATTENDANCE_EXPORT_TICK closed=%s materialized=%s processed=%s",
            closed,
            materialized,
            len(results),
        )
    return results


def get_attendance_dashboard(platform_id: int, course_date: str, *, center_account_id: int | None = None) -> dict[str, Any]:
    platform = attendance_repo.get_accessible_platform(platform_id, center_account_id)
    if not platform:
        raise LookupError("Plateforme introuvable")
    course_session = attendance_repo.get_course_session_for_date(platform_id, course_date)
    participants: list[dict[str, Any]] = []
    if course_session:
        rows = attendance_repo.list_presence_logs_for_session(platform_id, int(course_session["id"]))
        participants = consolidate_presence(
            rows,
            scheduled_at=course_session["scheduled_at"],
            timezone_name=course_session.get("timezone") or "Europe/Paris",
        )
    roster = list_explicit_course_reminder_recipients(platform_id) if course_session else []
    roster_by_hash = {
        course_invitation_recipient_hash(item.get("email")): item for item in roster
    }
    present_hashes = set()
    students = []
    for participant in participants:
        participant_hash = participant["key"].removeprefix("invite:") if participant["key"].startswith("invite:") else ""
        enrolled = roster_by_hash.get(participant_hash)
        if enrolled:
            present_hashes.add(participant_hash)
        slots = [
            {"start": start.strftime("%H:%M:%S"), "end": end.strftime("%H:%M:%S")}
            for start, end in participant["intervals"]
        ]
        students.append({
            "id": f"recipient:{enrolled['id']}" if enrolled else participant["key"],
            "email": enrolled.get("email") if enrolled else participant["email"],
            "nom": enrolled.get("nom") if enrolled else participant["nom"],
            "prenom": enrolled.get("prenom") if enrolled else participant["prenom"],
            "attendance": {
                "course_date": course_date,
                "slots": slots,
                "total_minutes": participant["total_seconds"] // 60,
                "total_seconds": participant["total_seconds"],
                "status": "present" if participant["total_seconds"] > 0 else "absent",
                "source": "automatic",
            },
        })
    for recipient_hash, enrolled in roster_by_hash.items():
        if recipient_hash in present_hashes:
            continue
        students.append({
            "id": f"recipient:{enrolled['id']}",
            "email": enrolled.get("email") or "",
            "nom": enrolled.get("nom") or "",
            "prenom": enrolled.get("prenom") or "",
            "attendance": {
                "course_date": course_date,
                "slots": [],
                "total_minutes": 0,
                "total_seconds": 0,
                "status": "absent",
                "source": "automatic",
            },
        })
    students.sort(key=lambda item: (str(item.get("nom") or "").casefold(), str(item.get("prenom") or "").casefold()))
    exports = []
    for item in attendance_repo.list_daily_exports(
        platform_id,
        center_account_id=center_account_id,
    ):
        exports.append({
            "id": int(item["id"]),
            "center_account_id": int(item["center_account_id"]),
            "center_platform_number": int(item["center_platform_number"]),
            "course_session_id": int(item["course_session_id"]),
            "teacher_module_id": (
                int(item["teacher_module_id"])
                if item.get("teacher_module_id") is not None
                else None
            ),
            "course_date": item["course_date"].isoformat(),
            "status": item["status"],
            "filename": item.get("filename"),
            "size_bytes": int(item.get("size_bytes") or 0),
            "participant_count": int(item.get("participant_count") or 0),
            "generated_at": item["generated_at"].isoformat() if item.get("generated_at") else None,
            "available_at": item["available_at"].isoformat() if item.get("available_at") else None,
        })
    return {
        "success": True,
        "platform": {
            "id": int(platform["id"]),
            "number": int(platform["center_platform_number"]),
            "name": platform["name"],
            "center_account_id": int(platform["center_account_id"]),
            "teacher_module_id": (
                int(platform["source_module_id"])
                if platform.get("source_module_id") is not None
                else None
            ),
        },
        "course_date": course_date,
        "course_session": ({
            "id": int(course_session["id"]),
            "session_index": int(course_session.get("session_index") or 0),
            "scheduled_at": course_session["scheduled_at"].isoformat(),
            "status": course_session.get("status"),
        } if course_session else None),
        "students": students,
        "recent_dates": [
            {
                "course_date": item["course_date"],
                "student_count": item["participant_count"],
                "status": item["status"],
                "export_id": item["id"],
            }
            for item in exports[:20]
        ],
        "daily_exports": exports,
        "recent_weeks": [],
    }
