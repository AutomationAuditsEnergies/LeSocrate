# export_service.py - Service d'export Excel
from collections import defaultdict
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import tempfile
from utils.logger import get_logger

logger = get_logger(__name__)


def _format_seconds(total_seconds):
    """Formate un nombre de secondes en 'Xh Ymin Zsec' ou 'Xmin Zsec'"""
    total_seconds = int(total_seconds)
    heures = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    secondes = total_seconds % 60
    if heures > 0:
        return f"{heures}h {minutes}min {secondes}sec"
    return f"{minutes}min {secondes}sec"


def generate_excel_export(logs_data):
    """
    Génère un fichier Excel à partir des données de logs.
    Colonnes A-F : détail des sessions
    Colonnes H-J : récapitulatif temps total par utilisateur

    Args:
        logs_data: Liste de tuples (id, nom, prenom, arrivee, depart)

    Returns:
        Chemin du fichier temporaire Excel généré
    """
    try:
        logger.info(f"📊 Génération export Excel pour {len(logs_data)} lignes")

        wb = Workbook()
        ws = wb.active
        ws.title = "Logs"

        # --- En-têtes détail (A-F) ---
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        summary_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")

        for col, title in enumerate(["ID", "Nom", "Prénom", "Arrivée", "Départ", "Durée session"], start=1):
            cell = ws.cell(row=1, column=col, value=title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # --- Données détail + calcul totaux par utilisateur ---
        totaux_par_user = defaultdict(int)   # clé: (nom, prenom) → secondes totales

        now = datetime.now()

        for row in logs_data:
            id_, nom, prenom, arrivee, depart = row[0], row[1], row[2], row[3], row[4]
            dt1 = datetime.strptime(arrivee, "%Y-%m-%d %H:%M:%S")
            if depart:
                dt2 = datetime.strptime(depart, "%Y-%m-%d %H:%M:%S")
                seconds = (dt2 - dt1).total_seconds()
                minutes = int(seconds // 60)
                secondes = int(seconds % 60)
                duree = f"{minutes}min {secondes}sec"
            else:
                seconds = (now - dt1).total_seconds()
                minutes = int(seconds // 60)
                secondes = int(seconds % 60)
                duree = f"{minutes}min {secondes}sec (en cours)"
            totaux_par_user[(nom, prenom)] += seconds
            ws.append([id_, nom, prenom, arrivee, depart or "", duree])

        # --- En-têtes récapitulatif (colonnes H=8, I=9, J=10) ---
        for col, title in enumerate(["Nom", "Prénom", "Temps total de connexion"], start=8):
            cell = ws.cell(row=1, column=col, value=title)
            cell.font = header_font
            cell.fill = summary_fill
            cell.alignment = Alignment(horizontal="center")

        # --- Données récapitulatif, triées par nom ---
        for i, ((nom, prenom), total_sec) in enumerate(
            sorted(totaux_par_user.items(), key=lambda x: (x[0][0], x[0][1])),
            start=2
        ):
            ws.cell(row=i, column=8, value=nom)
            ws.cell(row=i, column=9, value=prenom)
            ws.cell(row=i, column=10, value=_format_seconds(total_sec))

        # --- Largeurs de colonnes ---
        column_widths = {1: 6, 2: 18, 3: 18, 4: 20, 5: 20, 6: 16, 8: 18, 9: 18, 10: 28}
        for col, width in column_widths.items():
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

        # --- Date d'export en bas du tableau principal ---
        last_row = ws.max_row + 2
        date_cell = ws.cell(row=last_row, column=1, value=f"Exporté le {datetime.now().strftime('%d/%m/%Y')}")
        date_cell.font = Font(italic=True, color="888888")

        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(tmp.name)
        tmp.seek(0)

        logger.info("✅ Export Excel généré avec succès")
        return tmp.name

    except Exception as e:
        logger.error(f"❌ Erreur génération Excel: {e}")
        raise


def _format_minutes(total_minutes):
    total_minutes = int(total_minutes or 0)
    heures = total_minutes // 60
    minutes = total_minutes % 60
    if heures and minutes:
        return f"{heures}h {minutes:02d}"
    if heures:
        return f"{heures}h"
    return f"{minutes}min"


def _format_slots(slots):
    if not slots:
        return ""
    return " ; ".join(
        f"{slot.get('start', '')}-{slot.get('end', '')}"
        for slot in slots
        if slot.get("start") or slot.get("end")
    )


def generate_attendance_excel_export(records, platform_name="Formation"):
    """
    Génère un export Excel des présences consolidées par journée de cours.

    Args:
        records: liste de dicts avec student, course_date, slots, total_minutes, status, notes.
        platform_name: nom affiché dans l'export.

    Returns:
        Chemin du fichier temporaire Excel généré.
    """
    try:
        logger.info(f"📊 Génération export présences pour {len(records)} lignes")
        wb = Workbook()
        ws = wb.active
        ws.title = "Présences"
        summary = wb.create_sheet("Récapitulatif")

        header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        summary_fill = PatternFill(start_color="4C1D95", end_color="4C1D95", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        ws.cell(row=1, column=1, value=f"Présences · {platform_name}")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)

        headers = ["Date du cours", "Nom", "Prénom", "Email", "Statut", "Créneaux", "Temps présent", "Minutes", "Notes"]
        for col, title in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col, value=title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        totals = defaultdict(int)
        latest = {}
        for row_idx, record in enumerate(records, start=4):
            key = (record.get("nom") or "", record.get("prenom") or "", record.get("email") or "")
            total_minutes = int(record.get("total_minutes") or 0)
            totals[key] += total_minutes
            latest[key] = record.get("course_date")
            ws.cell(row=row_idx, column=1, value=record.get("course_date"))
            ws.cell(row=row_idx, column=2, value=record.get("nom"))
            ws.cell(row=row_idx, column=3, value=record.get("prenom"))
            ws.cell(row=row_idx, column=4, value=record.get("email"))
            ws.cell(row=row_idx, column=5, value=record.get("status"))
            ws.cell(row=row_idx, column=6, value=_format_slots(record.get("slots") or []))
            ws.cell(row=row_idx, column=7, value=_format_minutes(total_minutes))
            ws.cell(row=row_idx, column=8, value=total_minutes)
            ws.cell(row=row_idx, column=9, value=record.get("notes") or "")

        for col, width in {1: 16, 2: 20, 3: 20, 4: 30, 5: 16, 6: 28, 7: 16, 8: 12, 9: 34}.items():
            ws.column_dimensions[ws.cell(row=3, column=col).column_letter].width = width

        for col, title in enumerate(["Nom", "Prénom", "Email", "Temps total", "Minutes totales", "Dernière date"], start=1):
            cell = summary.cell(row=1, column=col, value=title)
            cell.font = header_font
            cell.fill = summary_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, ((nom, prenom, email), total_minutes) in enumerate(
            sorted(totals.items(), key=lambda item: (item[0][0], item[0][1], item[0][2])),
            start=2,
        ):
            summary.cell(row=row_idx, column=1, value=nom)
            summary.cell(row=row_idx, column=2, value=prenom)
            summary.cell(row=row_idx, column=3, value=email)
            summary.cell(row=row_idx, column=4, value=_format_minutes(total_minutes))
            summary.cell(row=row_idx, column=5, value=total_minutes)
            summary.cell(row=row_idx, column=6, value=latest.get((nom, prenom, email)) or "")

        for col, width in {1: 20, 2: 20, 3: 30, 4: 18, 5: 16, 6: 16}.items():
            summary.column_dimensions[summary.cell(row=1, column=col).column_letter].width = width

        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(tmp.name)
        tmp.seek(0)
        logger.info("✅ Export présences généré avec succès")
        return tmp.name
    except Exception as e:
        logger.error(f"❌ Erreur génération export présences: {e}")
        raise
